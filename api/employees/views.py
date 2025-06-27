import openpyxl.utils
import openpyxl.utils.datetime
from rest_framework import viewsets, status
from .models import Employee, EmploymentRecord
from .serializers import EmployeeSerializer, EmployeeRecordSerializer
from rest_framework.decorators import action
from rest_framework.response import Response
import io
import csv
from rest_framework import filters
from rest_framework.permissions import IsAuthenticated
from core.permissions import IsCompanyUser, IsTalentVerifyAdmin
from api.companies.models import Department
import openpyxl

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'employee_id', 'employment_records__company__name', 
                     'employment_records__department__name', 'employment_records__role',
                     'employment_records__date_started', 'employment_records__date_left']

    def get_permissions(self):
        """
        - List/retrieve/search: Any authenticated user
        - Create/update/delete/bulk_create: Company users or Talent Verify admins
        """
        if self.action in ['list', 'retrieve', 'search']:
            permission_classes = [IsAuthenticated]
        elif self.action in ['create', 'update', 'destroy', 'bulk_create']:
            permission_classes = [IsCompanyUser | IsTalentVerifyAdmin]
        else:
            permission_classes = []
        return [permission() for permission in permission_classes]
    
    # def get_queryset(self):
    #     """
    #     Company users can only see employees with records at their company.
    #     Talent Verify admins can see all employees.
    #     """
    #     user = self.request.user
    #     if user.user_type == 'admin':
    #         return Employee.objects.all()
    #     elif user.user_type == 'company':
    #         return Employee.objects.filter(employment_records__company=user.company).distinct()
    #     return Employee.objects.all()

    @action(detail=False, methods=['get'])
    def search(self, request):
         
         
        queryset = self.get_queryset()

        # Get search parameters
        name = request.query_params.get('name', None)
        employer = request.query_params.get('employer', None)
        position = request.query_params.get('position', None)
        department = request.query_params.get('department', None)
        year_started = request.query_params.get('year_started', None)
        year_left = request.query_params.get('year_left', None)

        # Apply filters based on search parameters
        if name:
            queryset = queryset.filter(name__icontains=name)
        if employer:
            queryset = queryset.filter(employment_records__company__name__icontains=employer)
        if position:
            queryset = queryset.filter(employment_records__role__icontains=position)
        if department:
            queryset = queryset.filter(employment_records__department__name__icontains=department)
        if year_started:
            queryset = queryset.filter(employment_records__date_started__year=year_started)
        if year_left:
            queryset = queryset.filter(employment_records__date_left__year=year_left)

        queryset = queryset.distinct()

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        
        file = request.FILES.get('file')
        company_id = request.data.get('company_id')
        required_fields = ['employee_name', 'department_name', 'role', 'date_started']

        if not company_id:
            return Response({"error": "Company ID is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        file_name = file.name.lower()
        if file_name.endswith('.csv') or file_name.endswith('.txt'):
            try:
                decoded_file = file.read().decode('utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                decoded_file = file.read().decode('latin1')
            io_string = io.StringIO(decoded_file)

            reader = csv.DictReader(io_string)

            missing_fields = [field for field in required_fields if field not in reader.fieldnames]

            if missing_fields:
                return Response({"error": f"CSV or text file is missing required fields: {', '.join(missing_fields)}"}, 
                                status=status.HTTP_400_BAD_REQUEST)
            
        elif file_name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1,max_row=1))]

            missing_fields = [field for field in required_fields if field not in headers]

            if missing_fields:
                return Response({"error": f"excel file is missing required fields: {', '.join(missing_fields)}"}, 
                                status=status.HTTP_400_BAD_REQUEST)
            
            reader = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers,row))
                for date_field in ['date_started','date_left']:
                    value = row_dict.get(date_field)
                    if isinstance(value,(int,float)):
                        try:
                            row_dict[date_field] = openpyxl.utils.datetime.from_excel(value).date()
                        except Exception:
                            row_dict[date_field] = None
                reader.append(row_dict)
        else:
            return Response({'error':'Unsupported file format'},status=status.HTTP_400_BAD_REQUEST)

        emp_created_count = 0
        rec_created_count = 0
        emp_updated_count = 0
        rec_updated_count = 0
        errors = []

        for row in reader:
            try:
                # Create or update employee
                employee, created = Employee.objects.get_or_create(
                    name=row['employee_name'],
                    defaults={
                        'employee_id': row.get('employee_id', None),
                    }
                )
                if created:
                    emp_created_count += 1
                else:
                    emp_updated_count += 1
                    
                # Get or create department
                department, _ = Department.objects.get_or_create(
                    company_id=company_id,
                    name=row['department_name']
                )
                
                existing_record = EmploymentRecord.objects.filter(employee=employee,
                                                                  company_id=company_id,
                                                                  department=department,
                                                                  role=row['role'],
                                                                  date_started=row['date_started'],
                                                                  duties=row['duties']).first()
                if existing_record:
                    if not existing_record.date_left and row['date_left']:
                        print(existing_record)
                        existing_record.date_left=row['date_left']
                        existing_record.save()
                        rec_updated_count+=1
                else:
                    # Create employment record
                    employment_record= EmploymentRecord.objects.create(
                        employee=employee,
                        company_id=company_id,
                        department=department,
                        role=row['role'],
                        date_started = row['date_started'],
                        date_left = row.get('date_left') or None,
                        duties = row.get('duties', ''),
                    )
                    rec_created_count+=1

            except Exception as e:
                errors.append(f"Error processing row {row}: {str(e)}")

        return Response({
            'emp_created_count': emp_created_count,
            'rec_created_count': rec_created_count,
            'emp_updated_count': emp_updated_count,
            'rec_updated_count': rec_updated_count,
            'errors': errors
        })
    
class EmploymentRecordViewSet(viewsets.ModelViewSet):

    serializer_class = EmployeeRecordSerializer

    def get_queryset(self):
        employee_id = self.kwargs['employee_pk']
        return EmploymentRecord.objects.filter(employee_id=employee_id)
    
    def perform_create(self, serializer):
        employee_id = self.kwargs['employee_pk']
        serializer.save(employee_id=employee_id)

    